"""
Tender Breakdown Autofill Engine
Reads cash sheet workbooks and fills a master Tender Breakdown spreadsheet.
"""

import os
import re
from pathlib import Path
from openpyxl import load_workbook
from dateutil.parser import parse as dateutil_parse

try:
    from .config import (
        FILENAME_TO_MASTER_NAME,
        IMPORTANT_CASHEET_DATA_COL,
        LOCATION_START_COL,
        FIRST_DATE_ROW,
        DATE_COL,
    )
except ImportError:
    from config import (
        FILENAME_TO_MASTER_NAME,
        IMPORTANT_CASHEET_DATA_COL,
        LOCATION_START_COL,
        FIRST_DATE_ROW,
        DATE_COL,
    )


# ═══════════════════════════════════════════════════════════════
#  PROCESSING TRACKER
# ═══════════════════════════════════════════════════════════════

class ProcessingTracker:
    """Tracks successes, failures, and warnings across all files."""

    def __init__(self, on_event=None):
        self.successful = []
        self.failed = []
        self.warnings = []
        self._on_event = on_event

    def _emit(self, kind, msg):
        if self._on_event:
            self._on_event(kind, msg)

    def log(self, msg):
        self._emit("info", msg)

    def add_success(self, location, date, filename):
        self.successful.append((location, date, filename))
        self._emit("success", f"  ✅ {location} | {date} | {filename}")

    def add_failure(self, location, date, filename, error):
        self.failed.append((location, date, filename, error))
        self._emit(
            "error", f"  ❌ {location} | {date} | {filename}\n     └─ {error}")

    def add_warning(self, msg):
        self.warnings.append(msg)
        self._emit("warning", f"  ⚠️  {msg}")

    def print_summary(self):
        total = len(self.successful) + len(self.failed)
        self.log(f"\n{'='*80}")
        self.log("📋 TENDER BREAKDOWN SUMMARY")
        self.log(f"{'='*80}")
        self.log(f"✅ Successful: {len(self.successful)} / {total}")
        for loc, date, fname in self.successful:
            self.log(f"   • {loc:<20} | {date} | {fname}")
        if self.failed:
            self.log(f"\n❌ Errors: {len(self.failed)}")
            for loc, date, fname, err in self.failed:
                d_str = str(date) if date else "N/A"
                self.log(
                    f"   • {loc or 'Unknown':<20} | {d_str:<10} | {fname}")
                self.log(f"     └─ {err}")
        self.log(f"{'='*80}\n")


# ═══════════════════════════════════════════════════════════════
#  CASHEET HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def get_casheet_date(ws):
    """
    Find the report date in a cash sheet worksheet.

    Strategy:
        1. Check known location (row 1, DATE_COL) directly.
        2. Fallback: scan first 5 rows for a 'Date:' label.

    Returns:
        datetime.date or None
    """
    # Fast path: known date column
    direct_val = ws.cell(row=1, column=DATE_COL).value
    if direct_val:
        try:
            return dateutil_parse(str(direct_val), fuzzy=True).date()
        except (ValueError, TypeError):
            pass

    # Fallback: scan for "Date:" label
    for row in ws.iter_rows(min_row=1, max_row=5):
        for i in range(len(row) - 1):
            cell_val = str(row[i].value).strip() if row[i].value else ""
            if "date:" in cell_val.lower():
                date_value = row[i + 1].value
                if date_value:
                    try:
                        return dateutil_parse(str(date_value), fuzzy=True).date()
                    except (ValueError, TypeError):
                        continue
    return None


def get_casheet_data(ws, keyword):
    """
    Search column B for a keyword, then extract the configured columns.

    Args:
        ws: worksheet object
        keyword: string to search for (case-insensitive, partial match)

    Returns:
        list of values, or empty list if not found
    """
    for row_idx in range(4, 60):
        cell_val = ws.cell(row=row_idx, column=2).value  # Column B
        if cell_val and keyword.lower() in str(cell_val).lower():
            data = []
            for col_idx in IMPORTANT_CASHEET_DATA_COL:
                val = ws.cell(row=row_idx, column=col_idx).value
                data.append(val if val is not None else 0)
            return data
    return []


def get_heritage_data(ws, keywords):
    """
    Special case: sum data from two register rows (Heritage/PHC).

    Args:
        ws: worksheet object
        keywords: list of 2 keyword strings

    Returns:
        list of summed values
    """
    data1 = get_casheet_data(ws, keywords[0]) or []
    data2 = get_casheet_data(ws, keywords[1]) or []

    if not data1 and data2:
        data1 = [0] * len(IMPORTANT_CASHEET_DATA_COL)
    if not data2:
        data2 = [0] * len(IMPORTANT_CASHEET_DATA_COL)

    return [a + b for a, b in zip(data1, data2)]


# ═══════════════════════════════════════════════════════════════
#  MASTER BREAKDOWN HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def find_master_row_by_date(ws, target_date):
    """
    Find the row in the master breakdown matching target_date.

    Strategy:
        1. Calculate expected row via date offset from first row.
        2. Verify the cell actually contains the target date.
        3. Fallback: scan all rows manually.
    """
    first_cell = ws.cell(row=FIRST_DATE_ROW, column=1).value
    if not first_cell:
        return _find_master_row_manual(ws, target_date)

    try:
        first_date = dateutil_parse(str(first_cell), fuzzy=True).date()
    except (ValueError, TypeError):
        return _find_master_row_manual(ws, target_date)

    date_diff = (target_date - first_date).days
    if date_diff < 0:
        return _find_master_row_manual(ws, target_date)

    expected_row = FIRST_DATE_ROW + date_diff
    cell_val = ws.cell(row=expected_row, column=1).value
    if cell_val is None:
        return _find_master_row_manual(ws, target_date)

    try:
        cell_date = dateutil_parse(str(cell_val), fuzzy=True).date()
    except (ValueError, TypeError):
        return _find_master_row_manual(ws, target_date)

    if cell_date != target_date:
        return _find_master_row_manual(ws, target_date)

    return expected_row


def _find_master_row_manual(ws, target_date):
    """Scan column A manually for the target date."""
    for row in ws.iter_rows(min_row=FIRST_DATE_ROW, max_col=1):
        cell_val = row[0].value
        if cell_val:
            try:
                cell_date = dateutil_parse(str(cell_val), fuzzy=True).date()
                if cell_date == target_date:
                    return row[0].row
            except (ValueError, TypeError):
                continue
    return None


def fill_tender_row(ws, data_values, row_idx, start_col):
    """Write a list of values into consecutive columns starting at start_col."""
    for i, value in enumerate(data_values):
        ws.cell(row=row_idx, column=start_col + i, value=value)


# ═══════════════════════════════════════════════════════════════
#  MAIN ENGINE
# ═══════════════════════════════════════════════════════════════

class TenderBreakdownEngine:
    """
    Processes cash sheet workbooks and fills the master Tender Breakdown.

    Usage:
        engine = TenderBreakdownEngine(
            casheet_dir="path/to/cash_sheets",
            master_path="path/to/Tender Breakdown.xlsx",
            on_event=callback,
            stop_event=threading_event,
        )
        engine.execute()
    """

    def __init__(self, casheet_dir, master_path, on_event=None, stop_event=None):
        self.casheet_dir = casheet_dir
        self.master_path = master_path
        self.tracker = ProcessingTracker(on_event=on_event)
        self.stop_event = stop_event

    def _stopped(self):
        return self.stop_event and self.stop_event.is_set()

    def execute(self):
        """Run the full tender breakdown autofill process."""

        # ── Validate paths ────────────────────────────────────────
        if not os.path.isdir(self.casheet_dir):
            self.tracker.log(
                f"❌ Cash sheets directory not found: {self.casheet_dir}")
            return self.tracker

        if not os.path.isfile(self.master_path):
            self.tracker.log(
                f"❌ Master Breakdown file not found: {self.master_path}")
            return self.tracker

        # ── Find cash sheet files ─────────────────────────────────
        files = [
            f for f in Path(self.casheet_dir).glob("*.xlsx")
            if not f.name.startswith("~$")
        ]

        if not files:
            self.tracker.log(
                "⚠️ No .xlsx files found in cash sheets directory.")
            return self.tracker

        self.tracker.log(
            f"📂 Found {len(files)} cash sheet file(s). Starting...")

        # ── Load Master Breakdown ONCE ────────────────────────────
        try:
            wb_master = load_workbook(self.master_path)
            ws_master = wb_master.active
        except PermissionError:
            self.tracker.log(
                "❌ Cannot open Master Breakdown — please close the file and try again.")
            return self.tracker
        except Exception as e:
            self.tracker.log(f"❌ Error loading Master Breakdown: {e}")
            return self.tracker

        # ── Process each cash sheet ───────────────────────────────
        for file_path in files:
            if self._stopped():
                self.tracker.log("🛑 Aborting — stopped by user.")
                break

            self._process_casheet(file_path, ws_master)

        # ── Save Master ONCE ──────────────────────────────────────
        if not self._stopped():
            try:
                wb_master.save(self.master_path)
                self.tracker.log("💾 Master Breakdown saved successfully!")
            except PermissionError:
                self.tracker.log(
                    "❌ Could not save — please close 'Tender Breakdown.xlsx' and try again.")
            except Exception as e:
                self.tracker.log(f"❌ Error saving: {e}")

        wb_master.close()

        # ── Summary ───────────────────────────────────────────────
        self.tracker.print_summary()
        return self.tracker

    def _process_casheet(self, casheet_path, master_ws):
        """Process a single cash sheet workbook."""
        filename = casheet_path.name
        self.tracker.log(f"\n{'='*80}")
        self.tracker.log(f"📄 {filename}")

        # ── Match filename to config key ──────────────────────────
        clean_name = re.sub(
            r'\s*\d{1,2}-\d{1,2}-\d{2,4}.*', '', casheet_path.stem
        ).strip().lower()

        mappings = FILENAME_TO_MASTER_NAME.get(clean_name)
        if not mappings:
            self.tracker.add_failure(
                "Unknown", None, filename,
                f"No config mapping for filename key: '{clean_name}'")
            return

        # ── Open cash sheet workbook ──────────────────────────────
        try:
            wb = load_workbook(casheet_path, data_only=True)
        except Exception as e:
            self.tracker.add_failure("File Error", None, filename, str(e))
            return

        # ── Process each worksheet (day) ──────────────────────────
        for ws in wb.worksheets:
            if self._stopped():
                break

            report_date = get_casheet_date(ws)
            if not report_date:
                self.tracker.add_failure(
                    "General", ws.title, filename,
                    "Could not find a valid date in the worksheet header")
                continue

            master_row = find_master_row_by_date(master_ws, report_date)
            if not master_row:
                self.tracker.add_failure(
                    "General", str(report_date), filename,
                    "Date not found in Master Breakdown")
                continue

            self.tracker.log(
                f"  📅 {ws.title} → {report_date} → Master row {master_row}")

            # ── Fill each mapped location ─────────────────────────
            for mapping_dict in mappings:
                for master_loc_name, casheet_keyword in mapping_dict.items():

                    start_col = LOCATION_START_COL.get(master_loc_name.lower())
                    if not start_col:
                        self.tracker.add_failure(
                            master_loc_name, str(report_date), filename,
                            f"No start column configured for '{master_loc_name}'")
                        continue

                    # Extract data (special case for PHC/Heritage)
                    if master_loc_name.lower() == "phc":
                        data = get_heritage_data(ws, casheet_keyword)
                    else:
                        data = get_casheet_data(ws, casheet_keyword)

                    if data:
                        fill_tender_row(master_ws, data, master_row, start_col)
                        self.tracker.add_success(
                            master_loc_name, str(report_date), filename)
                    else:
                        kw_display = casheet_keyword if isinstance(
                            casheet_keyword, str) else str(casheet_keyword)
                        self.tracker.add_failure(
                            master_loc_name, str(report_date), filename,
                            f"Keyword '{kw_display}' not found in worksheet rows")

        wb.close()
