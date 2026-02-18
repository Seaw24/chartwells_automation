"""
Grubhub Autofill Verification Script
Parses the Grubhub CSV and reads back the cash sheets to verify
that data was written correctly. Compares expected vs actual values.
"""

from config import (
    GRUBHUB_VENUE_MAP, GRUBHUB_TENDERS, CASHEET_TENDERS,
    FILL_COL_MAP, CASH_SHEET_FOLDER, REPORTS_FOLDER,
)
import os
import sys
import csv
from datetime import datetime
from openpyxl import load_workbook

# Add parent dir so we can import config
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))

# Tender columns to verify
TENDER_COLS = {
    "contract_card": FILL_COL_MAP.get("contract_card"),
    "flex": FILL_COL_MAP.get("flex"),
    "transfer": FILL_COL_MAP.get("transfer"),
    "ucash": FILL_COL_MAP.get("ucash"),
    "dining_dollars": FILL_COL_MAP.get("dining_dollars"),
    "visa": FILL_COL_MAP.get("visa"),
    "amex": FILL_COL_MAP.get("amex"),
    "discover": FILL_COL_MAP.get("discover"),
    "mc": FILL_COL_MAP.get("mc"),
    "coupons": FILL_COL_MAP.get("coupons"),
    "ushop": FILL_COL_MAP.get("ushop"),
    "chartwellsDCB": FILL_COL_MAP.get("chartwellsDCB"),
}

COUNT_COL = FILL_COL_MAP.get("count")
TOTAL_SALES_COL = FILL_COL_MAP.get("total_sales")
LOCATION_COL = FILL_COL_MAP.get("location")


def parse_grubhub_expected(grubhub_path):
    """
    Parse the Grubhub CSV and compute expected values per (casheet_file, location, weekday).

    Returns:
        dict: {(casheet_pattern, location, weekday): {count, total_sales, tenders: {...}}}
    """
    expected = {}

    with open(grubhub_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # skip header

        for row in reader:
            if len(row) < 17 or row[0].strip() in ("Total", ""):
                continue

            date_str = row[0].strip()
            # Parse date
            date_obj = None
            for fmt in ("%m-%d-%y", "%m/%d/%Y", "%m-%d-%Y"):
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    break
                except ValueError:
                    continue
            if date_obj is None:
                continue

            weekday = date_obj.strftime("%A")
            venue = row[3].strip()

            # Handle meal transfer suffix
            if venue.endswith(" - Meal Transfer"):
                venue = venue[:-len(" - Meal Transfer")].strip()

            if venue not in GRUBHUB_VENUE_MAP:
                continue

            casheet_pattern, location = GRUBHUB_VENUE_MAP[venue]
            payment_method = row[5].strip()

            try:
                tax = float(row[8].strip().replace("$", "").replace(",", ""))
                sale = float(row[9].strip().replace("$", "").replace(",", ""))
                meal_count = int(row[16].strip().replace(",", ""))
            except (ValueError, IndexError):
                continue

            key = (casheet_pattern, location, weekday)
            if key not in expected:
                expected[key] = {
                    "count": 0,
                    "total_sales": 0.0,
                    "tenders": {k: 0.0 for k in CASHEET_TENDERS},
                }

            expected[key]["count"] += meal_count
            expected[key]["total_sales"] += sale

            if payment_method in GRUBHUB_TENDERS:
                tender_key = GRUBHUB_TENDERS[payment_method]
                expected[key]["tenders"][tender_key] += sale

    return expected


def find_casheet_file(pattern, files):
    for f in files:
        if pattern.strip() in f.strip():
            return f
    return None


def verify():
    # Find Grubhub file
    grubhub_file = None
    for f in os.listdir(REPORTS_FOLDER):
        if f.startswith("SalesbyVenue") and f.endswith(".csv"):
            grubhub_file = os.path.join(REPORTS_FOLDER, f)
            break

    if not grubhub_file:
        print("No Grubhub CSV found in reports folder.")
        return

    print(f"Grubhub file: {os.path.basename(grubhub_file)}")
    print(f"Cash sheets : {CASH_SHEET_FOLDER}\n")

    expected = parse_grubhub_expected(grubhub_file)
    casheet_files = os.listdir(CASH_SHEET_FOLDER)

    # Cache open workbooks
    wb_cache = {}
    total_checks = 0
    passed = 0
    failed = 0
    skipped = 0

    for (casheet_pattern, location, weekday), exp in sorted(expected.items()):
        casheet_file = find_casheet_file(casheet_pattern, casheet_files)
        if casheet_file is None:
            print(
                f"  SKIP  {casheet_pattern}/{location}/{weekday} — no file found")
            skipped += 1
            continue

        filepath = os.path.join(CASH_SHEET_FOLDER, casheet_file)

        # Open workbook (cached)
        if filepath not in wb_cache:
            try:
                wb_cache[filepath] = load_workbook(filepath, data_only=True)
            except Exception as e:
                print(f"  SKIP  {casheet_file} — cannot open: {e}")
                skipped += 1
                continue

        wb = wb_cache[filepath]

        # Find sheet (case-insensitive)
        sheet_map = {s.lower(): s for s in wb.sheetnames}
        actual_sheet = sheet_map.get(weekday.lower())
        if actual_sheet is None:
            print(f"  SKIP  {casheet_file}/{weekday} — sheet not found")
            skipped += 1
            continue

        ws = wb[actual_sheet]

        # Find row
        target_row = None
        for r in range(4, ws.max_row + 1):
            val = ws.cell(r, LOCATION_COL).value
            if val and val.strip().lower() == location.lower():
                target_row = r
                break

        if target_row is None:
            print(
                f"  SKIP  {casheet_file}/{actual_sheet} — '{location}' not found")
            skipped += 1
            continue

        # Compare values
        total_checks += 1
        issues = []

        # Check count
        actual_count = ws.cell(target_row, COUNT_COL).value or 0
        if actual_count != exp["count"]:
            issues.append(
                f"count: expected {exp['count']}, got {actual_count}")

        # Check total_sales
        actual_sales = ws.cell(target_row, TOTAL_SALES_COL).value or 0.0
        if abs(float(actual_sales) - exp["total_sales"]) > 0.01:
            issues.append(
                f"total_sales: expected ${exp['total_sales']:.2f}, got ${float(actual_sales):.2f}")

        # Check each tender
        for tender_name, col in TENDER_COLS.items():
            if col is None:
                continue
            exp_val = exp["tenders"].get(tender_name, 0.0)
            actual_val = ws.cell(target_row, col).value
            actual_val = float(actual_val) if actual_val else 0.0

            if abs(actual_val - exp_val) > 0.01:
                issues.append(
                    f"{tender_name}: expected ${exp_val:.2f}, got ${actual_val:.2f}")

        if issues:
            failed += 1
            print(
                f"  FAIL  {casheet_file} / {actual_sheet} / row {target_row} ({location})")
            for issue in issues:
                print(f"        {issue}")
        else:
            passed += 1
            print(
                f"  OK    {casheet_file} / {actual_sheet} / row {target_row} ({location})")

    # Close workbooks
    for wb in wb_cache.values():
        wb.close()

    print(f"\n{'='*60}")
    print(f"VERIFICATION SUMMARY")
    print(f"  Passed  : {passed}")
    print(f"  Failed  : {failed}")
    print(f"  Skipped : {skipped}")
    print(f"  Total   : {total_checks + skipped}")
    print(f"{'='*60}")


if __name__ == "__main__":
    verify()
