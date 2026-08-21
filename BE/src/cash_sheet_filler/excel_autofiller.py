"""
Excel Autofiller Module
Handles automated filling of cash sheet Excel workbooks with parsed sales data.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import traceback
from .config import FILL_COL_MAP, CHECKING_COL_MAP, VERBOSE_TRACE
from ..utils import strip_accents


class ExcelAutofiller:
    """
    Automates the process of filling cash sheet Excel files with sales data.

    Attributes:
        xl_path (str): Path to the Excel workbook file
        location (str): Location name to find and fill in the worksheet
        week_day (str): Name of the worksheet tab (day of week)
        start_row (int): Starting row for location search (default: 4)
        row (int): Current row being processed
        wb: Openpyxl workbook object
        ws: Openpyxl worksheet object
    """

    def __init__(self, xl_path, location, week_day, tracker=None):
        """
        Initialize the ExcelAutofiller.

        Args:
            xl_path (str): Full path to the Excel workbook
            location (str): Location name to search for in the workbook
            week_day (str): Worksheet name (e.g., 'Monday', 'Tuesday')
        """
        self.xl_path = xl_path
        self.location = location
        self.week_day = week_day
        self.start_row = 4  # Data starts at row 4 (after headers)
        self.row = 0
        self.wb = None
        self.ws = None
        self.tracker = tracker
        # Set in save(): the over/short formula text, captured before the
        # workbook round-trip drops its cached result. See _capture_over_formula.
        self._over_formula = None
        self._over_verified = False

    # ─── Logging Helpers ─────────────────────────────────────────────────

    def _log(self, msg):
        """Standard info logging."""
        if self.tracker:
            self.tracker.log(msg)
        else:
            print(msg)

    def _log_error(self, msg):
        """Log filling errors with a visual indicator."""
        self._log(f"  ❌ {msg}")

    def _log_warning(self, msg):
        """Log filling warnings with a visual indicator."""
        self._log(f"  ⚠️  {msg}")

    def _trace(self, msg):
        """Verbose step-by-step line. Silent unless ``verbose_trace`` is on."""
        if self.tracker and hasattr(self.tracker, "trace"):
            self.tracker.trace(msg)
        elif VERBOSE_TRACE:
            print(f"     {msg}")

    # ─── Core Methods ────────────────────────────────────────────────────

    def open_workbook(self):
        """
        Open and load the Excel workbook and specified worksheet.

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.wb = load_workbook(self.xl_path)

            # Check if the worksheet exists (case-insensitive)
            sheet_map = {s.lower(): s for s in self.wb.sheetnames}
            actual_name = sheet_map.get(self.week_day.lower())
            if actual_name is None:
                self._log_error(
                    f"Worksheet '{self.week_day}' not found in workbook")
                return False

            self.ws = self.wb[actual_name]
            return True

        except FileNotFoundError:
            self._log_error(f"Workbook file not found: {self.xl_path}")
            return False

        except PermissionError:
            self._log_error(
                f"Permission denied: {self.xl_path}. Please close the file if it's currently open.")
            return False

        except Exception as e:
            self._log_error(f"Error opening workbook: {e}")
            traceback.print_exc()
            return False

    def find_row(self):
        location_col = FILL_COL_MAP.get("location")
        if location_col is None:
            self._log_error("'location' column not defined in FILL_COL_MAP")
            return False

        target = strip_accents(self.location.strip().lower())

        # Pass 1: exact match
        for r in range(self.start_row, self.ws.max_row + 1):
            cell_value = self.ws.cell(r, location_col).value
            if cell_value and strip_accents(cell_value.strip().lower()) == target:
                self.row = r
                self._log(
                    f"  ✓ Found location '{self.location}' at row {self.row}")
                return True

        # Pass 2: contains fallback (e.g. "grubhub" in "City Edge - grubhub")
        for r in range(self.start_row, self.ws.max_row + 1):
            cell_value = self.ws.cell(r, location_col).value
            if cell_value and target in strip_accents(cell_value.strip().lower()):
                self.row = r
                self._log(
                    f"  ✓ Found location '{self.location}' via partial match at row {self.row}")
                return True

        self._log_error(
            f"Location '{self.location}' not found in column {location_col}")
        return False

    def _capture_over_formula(self):
        """
        Record the over/short cell's formula text before the workbook is saved.

        openpyxl writes formulas but never evaluates them, and saving a
        workbook that was loaded with ``data_only=False`` discards the cached
        results Excel had computed. So by the time ``checking_tenders`` reads
        the cell back through a ``data_only=True`` reload, a formula cell is
        ``None`` — indistinguishable from a genuinely balanced zero. Knowing
        the cell held a formula lets us report "not verified" instead of
        claiming a validation that never happened.

        Returns the formula string, or None if the cell is not a formula.
        """
        over_col = CHECKING_COL_MAP.get("over")
        if over_col is None or self.ws is None or not self.row:
            return None
        raw = self.ws.cell(self.row, over_col).value
        return raw if isinstance(raw, str) and raw.startswith("=") else None

    def checking_tenders(self):
        """
        Verify tender calculations via the 'over/short' column.

        Sets ``self._over_verified`` to record whether the check was actually
        conclusive — a formula cell cannot be evaluated in-process, so it is
        reported as unverified rather than silently passing.

        Returns:
            bool: False only when over/short is provably non-zero or unreadable.
                  A cell we cannot evaluate returns True (nothing to fail on).
        """
        self._over_verified = False
        over_col = CHECKING_COL_MAP.get("over")

        if over_col is None:
            self._log_warning(
                "'over' column not defined in CHECKING_COL_MAP - skipping validation")
            return True

        over_value = self.ws.cell(self.row, over_col).value

        if over_value is None:
            # Empty because the cell is a formula we can't evaluate, or
            # genuinely empty. Either way we have not verified anything.
            if self._over_formula:
                self._log_warning(
                    f"{self.location}: over/short not verified — column "
                    f"{over_col} holds a formula that Excel recalculates "
                    "when the file is opened")
            else:
                self._log_warning(
                    f"{self.location}: over/short not verified — column "
                    f"{over_col} is empty")
            return True

        # Check if there's a discrepancy
        try:
            over_amount = float(over_value)
        except (ValueError, TypeError):
            self._log_error(
                f"Invalid value in 'over/short' column: {over_value}")
            return False

        if over_amount != 0:
            self._log_warning(
                f"Tender validation failed: 'over/short' is {over_amount}")
            return False

        self._over_verified = True
        return True

    def filling(self, parser):
        """
        Fill the worksheet with data from the parser dictionary.
        """
        try:
            # Step 1: Find the correct row for this location
            if not self.find_row():
                return False

            # Tax is typically on the row below the main data row
            tax_row = self.row + 1

            # Collected for the verbose trace so the reader can see exactly
            # which cell each number landed in, and which were cleared.
            written, cleared = [], []

            def put(r, c, value, label):
                self.ws.cell(r, c).value = value
                ref = f"{get_column_letter(c)}{r}"
                if value is None:
                    cleared.append(f"{ref} {label}")
                else:
                    written.append(f"{ref} {label} "
                                   + (f"{value:.2f}" if isinstance(value, float)
                                      else f"{value}"))

            # Step 2: Fill basic performance metrics
            date_col = FILL_COL_MAP.get("date")
            if date_col:
                put(1, date_col, parser.get("date"), "date")

            count_col = FILL_COL_MAP.get("count")
            if count_col and parser.get("count") is not None:
                put(self.row, count_col, parser.get("count"), "count")

            total_sales_col = FILL_COL_MAP.get("total_sales")
            if total_sales_col:
                put(self.row, total_sales_col,
                    parser.get("total_sales"), "sales")

            tax_col = FILL_COL_MAP.get("tax")
            if tax_col:
                # Same column as sales, one row down — that's the sheet layout.
                put(tax_row, tax_col, parser.get("tax"), "tax")

            # Step 3: Fill tender amounts
            tenders = parser.get("tenders", {})
            unmatched_tenders = []

            for tender_name, amount in tenders.items():
                if tender_name not in FILL_COL_MAP:
                    unmatched_tenders.append(tender_name)
                    continue

                col = FILL_COL_MAP[tender_name]

                # Write all non-zero amounts (including negatives); clear zeros
                if amount != 0:
                    put(self.row, col, amount, tender_name)
                else:
                    put(self.row, col, None, tender_name)

            # Step 4: sibling-venue figures. Swoop Swap counts go to their
            # own column (G). Choose Your Own Adventure sales go to the
            # "1,2,3" column (H), which the transfer tender also targets —
            # written after the tender loop so a zero transfer doesn't
            # clear it.
            swoop_col = FILL_COL_MAP.get("swoop_swap")
            if swoop_col and parser.get("swoop_swap") is not None:
                put(self.row, swoop_col, parser.get("swoop_swap"),
                    "swoop_swap")

            cyoa_col = FILL_COL_MAP.get("cyoa")
            if cyoa_col and parser.get("cyoa") is not None:
                put(self.row, cyoa_col, parser.get("cyoa"), "cyoa")

            # Report any unmatched tenders
            if unmatched_tenders:
                self._log_warning(
                    f"Unmatched tenders not filled: {', '.join(unmatched_tenders)}")

            if written:
                self._trace("│ wrote       : " + " · ".join(written))
            if cleared:
                # A zero tender is written as None on purpose — it erases last
                # run's value rather than leaving a stale number behind.
                self._trace("│ cleared (0) : " + " · ".join(cleared))

            self._log(f"  ✓ {self.location} filled successfully")
            return True

        except KeyError as e:
            self._log_error(f"Missing expected data key: {e}")
            return False

        except Exception as e:
            self._log_error(f"Error during filling: {e}")
            traceback.print_exc()
            return False

    def save(self):
        """
        Save the workbook and validate the filled data.

        Returns:
            bool: True if save successful and data validates correctly, False otherwise
        """
        # Calculate formula
        self.wb.calculation.calcMode = 'auto'
        self.wb.calculation.fullCalcOnLoad = True
        try:
            # Step 0: Note whether over/short is a formula — saving below
            # discards the cached value we'd need to evaluate it.
            self._over_formula = self._capture_over_formula()

            # Step 1: Save the workbook
            self.wb.save(self.xl_path)
            self._log(f"  💾 Saved: {self.location} casheet")

            # Step 2: Reload workbook with calculated formulas (data_only=True)
            self.wb = load_workbook(self.xl_path, data_only=True)
            sheet_map = {s.lower(): s for s in self.wb.sheetnames}
            actual_name = sheet_map.get(self.week_day.lower(), self.week_day)
            self.ws = self.wb[actual_name]

            # Step 3: Validate tender calculations
            is_correct = self.checking_tenders()

            if is_correct and self._over_verified:
                self._log(
                    f"  ✓ {self.location} casheet validated successfully")
            elif not is_correct:
                self._log_warning(
                    f"{self.location} validation warning - check over/short column")
            # The inconclusive case already logged its own reason in
            # checking_tenders — don't follow it with a "validated" line.

            return is_correct

        except PermissionError:
            self._log_error(
                f"Permission denied when saving: {self.xl_path}. Please close the file if it's currently open.")
            return False

        except Exception as e:
            self._log_error(f"Error saving workbook: {e}")
            traceback.print_exc()
            return False

    def close(self):
        """
        Properly close the workbook to free resources.
        """
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - ensures workbook is closed."""
        self.close()
