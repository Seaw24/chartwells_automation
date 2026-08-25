import os
from openpyxl import load_workbook

# Import your configuration
from config import CASH_SHEET_FOLDER, col_maps_for


def clear_all_cash_sheets():
    print(f"Looking for Cash Sheets in: {CASH_SHEET_FOLDER}\n")

    weekdays = ["Monday", "Tuesday", "Wednesday",
                "Thursday", "Friday", "Saturday", "Sunday"]

    # Loop through every file in the folder
    for filename in os.listdir(CASH_SHEET_FOLDER):
        if not filename.endswith(".xlsx") or filename.startswith("~"):
            continue

        filepath = os.path.join(CASH_SHEET_FOLDER, filename)
        print(f"Wiping {filename}...")

        # Columns vary per workbook (Kahlert Village's sheet is a column
        # short), so resolve them from the file name rather than assuming
        # the standard layout. Exclude 'location' so we keep the row names,
        # and 'date' so we don't wipe the header date cell in the data rows.
        fill_cols, _ = col_maps_for(filepath)
        cols_to_clear = set(
            col for key, col in fill_cols.items()
            if col is not None and key not in ["location", "date"]
        )

        try:
            wb = load_workbook(filepath)
            changed = False

            # Check every sheet tab in the workbook
            for sheet_name in wb.sheetnames:
                # If the tab is a day of the week, wipe its data
                if sheet_name.strip().title() in weekdays:
                    ws = wb[sheet_name]

                    # 1. Clear the date at the top (Row 1)
                    date_col = fill_cols.get("date")
                    if date_col:
                        ws.cell(row=1, column=date_col).value = None

                    # 2. Clear all the data rows (Row 4 down to the bottom)
                    # This naturally protects Rows 1, 2, and 3 (headers)
                    for row in range(5, ws.max_row + 1):
                        for col in cols_to_clear:
                            ws.cell(row=row, column=col).value = None

                    changed = True

            if changed:
                wb.save(filepath)
                print(f"  ✓ Wiped clean!")
            else:
                print(f"  ℹ️ No weekday tabs found.")

            wb.close()

        except PermissionError:
            print(f"  ❌ File is open in Excel! Please close it first.")
        except Exception as e:
            print(f"  ❌ Error: {e}")


if __name__ == "__main__":
    clear_all_cash_sheets()
    print("\n✅ All cash sheets are now blank and ready for the AutoFill engine!")
