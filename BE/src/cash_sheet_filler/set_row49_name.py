"""
Set Row 49 Location Name Script
Goes through every cash sheet Excel file and sets row 49, column 2
(the location name column) to "a" on all day-of-week worksheets.
"""

import os
from openpyxl import load_workbook

try:
    from .config import CASH_SHEET_FOLDER
except ImportError:
    from config import CASH_SHEET_FOLDER

ROW = 49
LOCATION_COL = 2  # Column B = location name
DAY_SHEETS = ["Friday", "Saturday", "Sunday", "Monday",
              "Tuesday", "Wednesday", "Thursday"]
NEW_NAME = "grubhub"


def main():
    files = [f for f in os.listdir(CASH_SHEET_FOLDER) if f.endswith(".xlsx")]

    if not files:
        print("No .xlsx files found in cash sheets folder.")
        return

    print(
        f"Setting row {ROW} location name to '{NEW_NAME}' in {len(files)} cash sheets...\n")

    for filename in sorted(files):
        filepath = os.path.join(CASH_SHEET_FOLDER, filename)

        try:
            wb = load_workbook(filepath)
        except Exception as e:
            print(f"  ❌ Could not open {filename}: {e}")
            continue

        updated_sheets = []

        for day in DAY_SHEETS:
            if day not in wb.sheetnames:
                continue

            ws = wb[day]
            old_value = ws.cell(ROW, LOCATION_COL).value
            ws.cell(ROW, LOCATION_COL).value = NEW_NAME
            updated_sheets.append(day)

        if updated_sheets:
            wb.save(filepath)
            old_val = old_value if old_value else "(empty)"
            print(
                f"  ✅ {filename:<30s} row {ROW} was '{old_val}' → now '{NEW_NAME}'  ({len(updated_sheets)} sheets)")
        else:
            print(f"  ⚠️  {filename:<30s} no day sheets found")

        wb.close()

    print("\nDone!")


if __name__ == "__main__":
    main()
